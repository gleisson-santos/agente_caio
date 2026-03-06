#!/usr/bin/env python3
"""
Gerador de planilhas Excel (.xlsx)
Uso: python generate_xlsx.py [arquivo_saida.xlsx]

Dependências: pip install openpyxl
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
import sys
import os

# === ESTILOS PADRÃO ===
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def criar_planilha(caminho_saida="out/planilha.xlsx"):
    """Cria uma planilha Excel de exemplo com dados e gráficos"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # === TÍTULO ===
    ws['A1'] = "Relatório de Vendas - 2024"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:E1')

    # === CABEÇALHOS ===
    headers = ["Mês", "Vendas", "Custos", "Lucro", "Margem %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # === DADOS ===
    dados = [
        ["Janeiro", 15000, 8000],
        ["Fevereiro", 18000, 9500],
        ["Março", 22000, 11000],
        ["Abril", 19500, 10200],
        ["Maio", 25000, 12500],
        ["Junho", 28000, 13800],
    ]

    for row_idx, linha in enumerate(dados, 4):
        # Mês
        ws.cell(row=row_idx, column=1, value=linha[0]).border = BORDER
        # Vendas
        cell_vendas = ws.cell(row=row_idx, column=2, value=linha[1])
        cell_vendas.number_format = '#,##0'
        cell_vendas.border = BORDER
        # Custos
        cell_custos = ws.cell(row=row_idx, column=3, value=linha[2])
        cell_custos.number_format = '#,##0'
        cell_custos.border = BORDER
        # Lucro (fórmula)
        ws.cell(row=row_idx, column=4, value=f'=B{row_idx}-C{row_idx}').border = BORDER
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        # Margem % (fórmula)
        ws.cell(row=row_idx, column=5, value=f'=D{row_idx}/B{row_idx}').border = BORDER
        ws.cell(row=row_idx, column=5).number_format = '0.0%'

    # === TOTAIS ===
    row_total = len(dados) + 4
    ws.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_total, column=1).fill = PatternFill("solid", fgColor="D9E2F3")
    for col in range(2, 5):
        cell = ws.cell(row=row_total, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row_total-1})')
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.number_format = '#,##0'
        cell.border = BORDER
    # Média da margem
    ws.cell(row=row_total, column=5, value=f'=AVERAGE(E4:E{row_total-1})')
    ws.cell(row=row_total, column=5).number_format = '0.0%'
    ws.cell(row=row_total, column=5).font = Font(bold=True)
    ws.cell(row=row_total, column=5).fill = PatternFill("solid", fgColor="D9E2F3")

    # === AJUSTAR LARGURA ===
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 14

    # === GRÁFICO DE BARRAS ===
    ws_graf = wb.create_sheet("Gráficos")

    # Copiar dados para gráfico
    ws_graf['A1'] = "Mês"
    ws_graf['B1'] = "Vendas"
    ws_graf['C1'] = "Custos"
    for i, linha in enumerate(dados):
        ws_graf.cell(row=i+2, column=1, value=linha[0])
        ws_graf.cell(row=i+2, column=2, value=linha[1])
        ws_graf.cell(row=i+2, column=3, value=linha[2])

    # Gráfico de barras
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Vendas vs Custos"
    chart1.style = 12
    chart1.y_axis.title = "Valor (R$)"
    chart1.x_axis.title = "Mês"
    chart1.width = 15
    chart1.height = 10

    data = Reference(ws_graf, min_col=2, min_row=1, max_col=3, max_row=7)
    cats = Reference(ws_graf, min_col=1, min_row=2, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws_graf.add_chart(chart1, "E2")

    # Gráfico de linha
    chart2 = LineChart()
    chart2.title = "Tendência de Vendas"
    chart2.style = 10
    chart2.y_axis.title = "Valor"
    chart2.width = 15
    chart2.height = 10

    data2 = Reference(ws_graf, min_col=2, min_row=1, max_row=7)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws_graf.add_chart(chart2, "E18")

    # Criar diretório se não existir
    os.makedirs(os.path.dirname(caminho_saida) if os.path.dirname(caminho_saida) else '.', exist_ok=True)

    wb.save(caminho_saida)
    print(f"Planilha criada: {caminho_saida}")
    return caminho_saida


# === FUNÇÕES AUXILIARES PARA USO COM LLMs ===

def novo_workbook():
    """Cria novo workbook"""
    return Workbook()

def adicionar_aba(wb, nome):
    """Adiciona nova aba ao workbook"""
    return wb.create_sheet(nome)

def definir_celula(ws, celula, valor, negrito=False, cor_fundo=None):
    """Define valor e estilo de uma célula"""
    ws[celula] = valor
    if negrito:
        ws[celula].font = Font(bold=True)
    if cor_fundo:
        ws[celula].fill = PatternFill("solid", fgColor=cor_fundo)
    return ws[celula]

def adicionar_tabela(ws, dados, linha_inicio=1, col_inicio=1, com_cabecalho=True):
    """
    Adiciona tabela de dados à planilha

    Args:
        ws: worksheet
        dados: lista de listas (primeira linha = cabeçalho se com_cabecalho=True)
        linha_inicio: linha inicial
        col_inicio: coluna inicial
        com_cabecalho: se True, primeira linha é formatada como cabeçalho
    """
    for i, linha in enumerate(dados):
        for j, valor in enumerate(linha):
            cell = ws.cell(row=linha_inicio + i, column=col_inicio + j, value=valor)
            cell.border = BORDER
            if i == 0 and com_cabecalho:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
    return ws

def adicionar_grafico_barras(ws, dados_ref, categorias_ref, posicao="E2", titulo="Gráfico"):
    """Adiciona gráfico de barras"""
    chart = BarChart()
    chart.type = "col"
    chart.title = titulo
    chart.style = 12
    chart.width = 15
    chart.height = 10
    chart.add_data(dados_ref, titles_from_data=True)
    chart.set_categories(categorias_ref)
    ws.add_chart(chart, posicao)
    return chart

def adicionar_grafico_pizza(ws, dados_ref, categorias_ref, posicao="E2", titulo="Gráfico"):
    """Adiciona gráfico de pizza"""
    chart = PieChart()
    chart.title = titulo
    chart.width = 12
    chart.height = 10
    chart.add_data(dados_ref, titles_from_data=True)
    chart.set_categories(categorias_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    ws.add_chart(chart, posicao)
    return chart

def adicionar_grafico_linha(ws, dados_ref, categorias_ref, posicao="E2", titulo="Gráfico"):
    """Adiciona gráfico de linha"""
    chart = LineChart()
    chart.title = titulo
    chart.style = 10
    chart.width = 15
    chart.height = 10
    chart.add_data(dados_ref, titles_from_data=True)
    chart.set_categories(categorias_ref)
    ws.add_chart(chart, posicao)
    return chart

def salvar_workbook(wb, caminho):
    """Salva o workbook"""
    os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else '.', exist_ok=True)
    wb.save(caminho)
    print(f"Arquivo salvo: {caminho}")


if __name__ == "__main__":
    saida = sys.argv[1] if len(sys.argv) > 1 else "out/planilha_exemplo.xlsx"
    criar_planilha(saida)
